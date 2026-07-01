#!/usr/bin/env python3
"""
Generador de Certificados / Reconocimientos / DC3 — AGASI
App local: sube Excel, elige plantilla, genera cientos de PDFs en segundos.
"""

import json
import os
import sys
import tempfile
import threading
import time
import webbrowser
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

from flask import (
    Flask,
    Response,
    jsonify,
    render_template,
    request,
    send_file,
    send_from_directory,
    stream_with_context,
)
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# Optimización: usar chromium_headless_shell (más liviano) en vez de chromium completo
# headless_shell es el motor de renderizado SIN interfaz gráfica. Como solo
# generamos PDFs (no navegación visual), headless_shell basta y ahorra ~70 MB.
# Esto se aplica ANTES de hacer sync_playwright() para que use el headless_shell
# que empaquetamos en el bundle.
import os as _os
_os.environ.setdefault("PLAYWRIGHT_CHROMIUM_USE_HEADLESS_NEW", "1")
# Solo forzamos PLAYWRIGHT_BROWSERS_PATH cuando es bundle (recursos extraídos a _MEIPASS).
# En dev mode dejamos que Playwright use su cache por defecto (~/.cache/ms-playwright)
# para que funcione aunque la carpeta local _ms-playwright esté vacía.
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    _os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", str(Path(sys._MEIPASS) / "_ms-playwright"))

# ─── Config ───
BASE_DIR = Path(__file__).parent.resolve()
# Cuando se ejecuta como binario empaquetado por PyInstaller, los recursos
# (plantillas, static, firmas) se extraen a sys._MEIPASS. Flask debe buscarlos ahí.
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    APP_ROOT = Path(sys._MEIPASS)
    # cwd o el directorio del ejecutable, para cosas que el usuario escribe
    # (output, logs) — lo que sí o sí va en el sistema de archivos del usuario
    RUNTIME_DIR = Path(sys.executable).parent.resolve()
else:
    APP_ROOT = BASE_DIR
    RUNTIME_DIR = BASE_DIR

PORT = 8765
HOST = "127.0.0.1"

app = Flask(__name__, template_folder=str(APP_ROOT / "templates"), static_folder=str(APP_ROOT / "static"))
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB max upload

# ─── Estado global del batch (en memoria, no DB) ───
batch_state = {
    "status": "idle",  # idle | running | done | error
    "total": 0,
    "completed": 0,
    "current_name": "",
    "started_at": None,
    "finished_at": None,
    "output_dir": "",
    "error": "",
}


# ─── Rutas principales ───


@app.route("/")
def index():
    """Página principal — subir Excel, elegir plantilla, mapear, previsualizar."""
    # Listar plantillas disponibles
    plantillas_dir = APP_ROOT / "plantillas"
    plantillas = []
    if plantillas_dir.exists():
        for f in sorted(plantillas_dir.glob("*.html")):
            plantillas.append(f.stem)
    return render_template("index.html", plantillas=plantillas)


@app.route("/api/plantilla/<nombre>")
def get_plantilla(nombre):
    """Devuelve el HTML de una plantilla."""
    path = APP_ROOT / "plantillas" / f"{nombre}.html"
    if not path.exists():
        return jsonify({"error": "Plantilla no encontrada"}), 404
    with open(path, "r", encoding="utf-8") as f:
        html = f.read()
    return jsonify({"html": html})


# ─── Plantillas Excel por tipo de documento ───
PLANTILLAS_EXCEL = {
    "reconocimiento": {
        "nombre": "Plantilla Reconocimiento",
        "headers": [
            "Nombre Completo",  # Apellido paterno, apellido materno, nombre(s)
            "Curso",
            "Duracion (hrs)",
            "Fecha",
            "Instructor",
            "Lugar",
            "Folio",  # Opcional — si va vacío se genera automáticamente
        ],
        "ejemplo": [
            "López Hernández María Fernanda",
            "Seguridad Industrial y Protección Civil",
            "40",
            "15/06/2026",
            "Ing. Roberto Martínez",
            "Pachuca, Hidalgo",
            "",  # Folio en blanco → auto-genera
        ],
    },
    "dc3": {
        "nombre": "Plantilla DC-3",
        # Orden basado en "quién llena cada campo":
        # 🟡 = lo llena el cliente del cliente (datos del participante + empresa)
        # 🟢 = lo llena AGASI (datos del curso + control interno)
        # Todos los amarillos van juntos primero, sin intercalar.
        "headers": [
            "Nombre",                    # 🟡 Apellido paterno, apellido materno, nombre(s)
            "CURP",                      # 🟡 18 caracteres
            "Puesto",                    # 🟡 Puesto del participante en la empresa
            "Empresa (Razón Social)",    # 🟡 Razón social completa
            "RFC (de la Empresa)",       # 🟡 12 caracteres - RFC de la empresa
            "Representante Legal",       # 🟡 Quien firma como representante de la empresa
            "Curso",                     # 🟢 Nombre del curso
            "Duracion (hrs)",            # 🟢 Horas
            "Fecha Inicio",              # 🟢 dd/mm/aaaa
            "Fecha Fin",                 # 🟢 dd/mm/aaaa
            "Area Tematica",             # 🟢 Código + nombre, ej: "6000 Seguridad"
            "Ocupacion",                 # 🟢 Catálogo Nacional de Ocupaciones
            "Instructor",                # 🟢 Nombre del instructor
            "Folio",                     # 🟢 Opcional — si va vacío se genera automáticamente
        ],
        "ejemplo": [
            "López Hernández María Fernanda",
            "LOHF900215HGPLRR09",
            "Supervisor",
            "Industrias Peoles",
            "IPE850101AB1",
            "Ing. Alejandro García Salinas",  # Representante Legal
            "Seguridad Industrial y Protección Civil",
            "40",
            "01/06/2026",
            "15/06/2026",
            "6000 Seguridad",
            "09.4 Protección",
            "Ing. Roberto Martínez",
            "",  # Folio en blanco → auto-genera
        ],
        # Columnas que se rellenan en amarillo en la plantilla (las primeras 6)
        "columnas_amarillas": 6,
    },
    "constancia": {
        "nombre": "Plantilla Constancia",
        "headers": [
            "Nombre Completo",  # Apellido paterno, apellido materno, nombre(s)
            "Curso",
            "Duracion (hrs)",
            "Fecha",
            "Instructor",
            "Folio",  # Opcional — si va vacío se genera automáticamente
        ],
        "ejemplo": [
            "Juan Pérez García",
            "Primeros Auxilios y RCP",
            "16",
            "17/06/2026",
            "Dra. Ana Ruiz",
            "",  # Folio en blanco → auto-genera
        ],
    },
}


@app.route("/api/plantilla-excel/<tipo>")
def download_plantilla_excel(tipo):
    """Genera y descarga una plantilla Excel con headers y fila de ejemplo."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if tipo not in PLANTILLAS_EXCEL:
        return jsonify(
            {
                "error": f"Tipo '{tipo}' no válido. Disponibles: {list(PLANTILLAS_EXCEL.keys())}"
            }
        ), 404

    info = PLANTILLAS_EXCEL[tipo]
    wb = Workbook()
    ws = wb.active
    ws.title = info["nombre"][:30]

    # ─── Estilos ───
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A4786")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    example_fill = PatternFill("solid", fgColor="F1F5F9")
    example_font = Font(name="Arial", size=10, color="64748B", italic=True)
    thin = Side(border_style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 🟡 Color amarillo para las celdas que llena el cliente del cliente
    yellow_fill = PatternFill("solid", fgColor="FFF3A0")  # amarillo suave, no agresivo
    # Color de la fuente para celdas amarillas: gris oscuro para que se lea
    yellow_font = Font(name="Arial", size=10, color="1F2937")

    # ─── Logo (esquina superior izquierda) ───
    # El logo está anclado en A1, pero NO ocupa ninguna columna — es una imagen
    # flotante que se superpone a la celda. La tabla de headers puede empezar
    # tranquilamente en columna A.
    logo_path = APP_ROOT / "AG_Principal.png"
    has_logo = logo_path.exists()
    # No se reserva ninguna columna para el logo (es solo visual)
    logo_width_cols = 0

    # ─── Fila 1: título del documento (a la derecha del logo) ───
    if has_logo:
        # El título se combina solo desde la columna después del logo hasta el final
        title_start_col = logo_width_cols + 1
    else:
        title_start_col = 1
    ws.merge_cells(
        start_row=1,
        start_column=title_start_col,
        end_row=1,
        end_column=len(info["headers"]),
    )
    title_cell = ws.cell(row=1, column=title_start_col, value=info["nombre"])
    title_cell.font = Font(name="Arial", size=16, bold=True, color="1A4786")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ─── Filas de instrucciones: dinámicamente según el número de líneas ───
    # El número exacto se calcula DESPUÉS de definir instr_lines (más abajo).
    # Por ahora usamos valores temporales que se sobreescribirán.
    instr_start_row = 2
    instr_start_col = logo_width_cols + 1 if has_logo else 1

    # placeholder; se sobreescribe después
    instr_end_row = 3
    header_row = 4
    example_row = 5

    # ─── Texto de instrucciones en formato lista ───
    if tipo == "dc3":
        instr_lines = [
            "INSTRUCCIONES:",
            "• Complete los datos de los participantes.",
            "• El campo NOMBRE debe ir como: Apellido Paterno + Apellido Materno + Nombre(s)   Ej: López Hernández María Fernanda",
            "• La fila de ejemplo debe borrarse antes de procesar.",
            "• La columna FOLIO es opcional — si la deja vacía, el sistema asigna una automáticamente.",
            "",
            "Para DC-3:",
            "• CURP debe tener exactamente 18 caracteres.",
            "• El RFC es el de la EMPRESA (no del participante). Persona Moral: 12 caracteres · Persona Física: 13 caracteres.",
            "• En EMPRESA escriba la Razón Social completa.",
            "• En ÁREA TEMÁTICA escriba: CÓDIGO + NOMBRE   Ej: 6000 Seguridad",
            "• REPRESENTANTE LEGAL puede quedar vacío — si lo deja vacío, se usa el valor por defecto configurado en el sistema.",
            "",
            "─── ¿QUIÉN LLENA CADA CAMPO? ───",
            "🟡  AMARILLO  →  Los llena el CLIENTE (usted):",
            "       Nombre, CURP, Puesto, Empresa, RFC, Representante Legal",
            "⬜  BLANCO    →  Los llena AGASI:",
            "       Curso, Duracion, Fechas, Area Tematica, Ocupacion, Instructor, Folio",
        ]
    elif tipo == "reconocimiento":
        instr_lines = [
            "INSTRUCCIONES:",
            "• Complete los datos de los participantes.",
            "• El campo NOMBRE debe ir como: Apellido Paterno + Apellido Materno + Nombre(s)   Ej: López Hernández María Fernanda",
            "• La fila de ejemplo debe borrarse antes de procesar.",
            "• La columna FOLIO es opcional — si la deja vacía, el sistema asigna una automáticamente.",
            "",
            "─── ¿QUIÉN LLENA CADA CAMPO? ───",
            "Todos los campos los llena el CLIENTE (usted).",
        ]
    elif tipo == "constancia":
        instr_lines = [
            "INSTRUCCIONES:",
            "• Complete los datos de los participantes.",
            "• El campo NOMBRE debe ir como: Apellido Paterno + Apellido Materno + Nombre(s)   Ej: Juan Pérez García",
            "• La fila de ejemplo debe borrarse antes de procesar.",
            "• La columna FOLIO es opcional — si la deja vacía, el sistema asigna una automáticamente.",
            "",
            "─── ¿QUIÉN LLENA CADA CAMPO? ───",
            "Todos los campos los llena el CLIENTE (usted).",
        ]
    else:
        instr_lines = ["INSTRUCCIONES: Complete los datos de los participantes."]

    # ─── Calcular el número de filas para instrucciones dinámicamente ───
    # DC-3 con la sección "¿quién llena cada campo?" tiene ~18 líneas.
    # Como las instrucciones ahora empiezan en columna B (no E), hay mucho
    # más ancho para el texto → caben más líneas por fila.
    n_lines = len(instr_lines)
    lines_per_row = 6  # ~6 líneas a 9pt caben en una fila (con instrucciones en col B)
    n_instr_rows = max(2, -(-n_lines // lines_per_row))  # ceil division, mínimo 2
    instr_end_row = instr_start_row + n_instr_rows - 1
    header_row = instr_end_row + 1
    example_row = header_row + 1

    # Ahora sí, mergear las celdas con el rango correcto
    ws.merge_cells(
        start_row=instr_start_row,
        start_column=instr_start_col,
        end_row=instr_end_row,
        end_column=len(info["headers"]),
    )

    # ─── Leyenda al inicio de las instrucciones (misma celda mergeada) ───
    # La leyenda va como primera línea, antes de "INSTRUCCIONES:", en la misma
    # celda mergeada. Una sola línea, no crea celdas nuevas.
    if tipo in ("dc3", "reconocimiento", "constancia"):
        # Solo la agregamos si no está ya al inicio
        if not instr_lines or not instr_lines[0].startswith("💡"):
            instr_lines.insert(0, "💡 Si no ves todo el texto, haz más grande la celda para ver todas las instrucciones completas.")
            instr_lines.insert(1, "")  # línea en blanco para separar

    instr_text = "\n".join(instr_lines)
    instr_cell = ws.cell(row=instr_start_row, column=instr_start_col, value=instr_text)
    # Instrucciones en NEGRITAS para que se lean bien (no en itálica)
    instr_cell.font = Font(name="Arial", size=9, bold=True, color="1F2937")
    instr_cell.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    for r in range(instr_start_row, instr_end_row + 1):
        ws.row_dimensions[r].height = 22

    # ─── Headers (con color amarillo en las primeras N columnas si aplica) ───
    n_amarillas = info.get("columnas_amarillas", 0) if tipo == "dc3" else 0

    # Si hay logo, asegurar que la fila del título tenga buena altura
    if has_logo:
        ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 30)

    # Headers desde la columna 1 (alineados a la izquierda, sin corrimiento)
    for col_idx, header in enumerate(info["headers"], 1):
        c = ws.cell(row=header_row, column=col_idx, value=header)
        c.font = header_font
        # El header siempre azul, el amarillo es para las celdas de datos
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
    ws.row_dimensions[header_row].height = 30

    # ─── Fila de ejemplo: con amarillo en las primeras N columnas ───
    for col_idx, val in enumerate(info["ejemplo"], 1):
        c = ws.cell(row=example_row, column=col_idx, value=val)
        if col_idx <= n_amarillas:
            # Celda amarilla: fuente oscura (no la del example_fill) para que se lea
            c.font = yellow_font
            c.fill = yellow_fill
        else:
            c.font = example_font
            c.fill = example_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border

    # ─── Filas vacías para estructura ───
    # Mantener el mismo color de fondo en celdas vacías para que se vea la "tabla" coloreada
    start_empty = example_row + 1
    end_empty = start_empty + 6  # 6 filas vacías para que el cliente capture
    for r in range(start_empty, end_empty):
        for col_idx in range(1, len(info["headers"]) + 1):
            c = ws.cell(row=r, column=col_idx, value="")
            if col_idx <= n_amarillas:
                c.fill = yellow_fill  # también amarillo en vacías
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # ─── Anchos de columna ───
    for col_idx, header in enumerate(info["headers"], 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header)),
            len(str(info["ejemplo"][col_idx - 1]))
            if col_idx <= len(info["ejemplo"])
            else 0,
            12,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 40)

    # ─── Insertar logo en la esquina superior izquierda ───
    if has_logo:
        try:
            img = XLImage(str(logo_path))
            # Logo chico: ~90 px de ancho, cabe en 1 columna sin forzarlo
            target_width = 90
            scale = target_width / img.width
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            # Anclar a la celda A1 (esquina superior izquierda) — se superpone
            # con la celda A1, pero como el logo es pequeño, no tapa nada
            img.anchor = "A1"
            ws.add_image(img)
        except Exception as e:
            print(f"[download_plantilla_excel] No se pudo insertar logo: {e}")

    # ─── Congelar panel ───
    # Congela desde la primera columna (A) y la fila de headers
    ws.freeze_panes = f"A{header_row + 1}"

    # ─── Configuración de impresión ───
    # Por defecto: orientación horizontal (landscape) para que quepan mejor las 14 columnas.
    # NO escalamos a "fit to 1 page" porque reduce demasiado la letra.
    # El usuario puede decidir si imprimir 1 o varias páginas desde Excel.
    from openpyxl.worksheet.page import PageMargins

    # Orientación horizontal (landscape)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    # Papel tamaño carta (Letter, común en México)
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    # Repetir la fila de headers en cada página al imprimir (útil si se imprime
    # en varias páginas, la primera fila de headers aparece en todas)
    ws.print_title_rows = f"{header_row}:{header_row}"
    # Márgenes normales (no mínimos, para que la impresión respire)
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75,
                                   header=0.3, footer=0.3)
    # Centrar horizontalmente
    ws.print_options.horizontalCentered = True

    # ─── Guardar ───
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"plantilla_{tipo}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/upload", methods=["POST"])
def upload_excel():
    """Recibe Excel, extrae filas y devuelve columnas + datos."""
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Archivo sin nombre"}), 400

    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        return jsonify({"error": "Solo Excel (.xlsx) o CSV"}), 400

    # Guardar temporalmente
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix)
    file.save(tmp.name)
    tmp.close()

    try:
        if file.filename.endswith(".csv"):
            import csv

            with open(tmp.name, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)
            example_row_indices = []
        else:
            wb = load_workbook(tmp.name, data_only=True)
            ws = wb.active
            rows = [
                [str(c.value) if c.value is not None else "" for c in row]
                for row in ws.iter_rows()
            ]

            # Detectar filas de ejemplo (tienen relleno gris F1F5F9 = RGB(241, 245, 249))
            example_row_indices = []
            for row_idx, row in enumerate(ws.iter_rows()):
                # Solo checar la primera celda con valor
                if row and row[0].value and row[0].fill and row[0].fill.fgColor:
                    rgb = row[0].fill.fgColor.rgb
                    if rgb and "F1F5F9" in str(rgb).upper():
                        example_row_indices.append(row_idx)
            wb.close()

        if len(rows) < 2:
            return jsonify(
                {
                    "error": "El Excel necesita una fila de encabezados y al menos 1 fila de datos"
                }
            ), 400

        # Detectar automáticamente la fila de encabezados
        # Detectar automáticamente la fila de encabezados
        # Busca la primera fila con texto en TODAS las celdas que parezcan nombres de campo
        # (Nombre, CURP, RFC, Folio, etc.) y que no sea un título o instrucciones
        palabras_clave = [
            "nombre",
            "curp",
            "rfc",
            "folio",
            "curso",
            "fecha",
            "duracion",
            "instructor",
            "lugar",
            "empresa",
            "puesto",
            "ocupacion",
            "area",
        ]

        header_idx = 0
        for idx, row in enumerate(rows[:30]):  # Buscar en las primeras 30 filas (DC-3 tiene ~18 líneas de instrucciones)
            row_text = " ".join(str(c).lower() for c in row if c)
            # Verificar que la fila tenga al menos 2 palabras clave (típico de headers)
            matches = sum(1 for kw in palabras_clave if kw in row_text)
            if matches >= 2:
                # Verificar que NO sea una fila de título/instrucciones (típicamente más cortas)
                non_empty = [c for c in row if c]
                if non_empty and all(len(str(c)) < 50 for c in non_empty):
                    header_idx = idx
                    break

        headers = rows[header_idx]
        data_rows = rows[header_idx + 1 :]

        # Filtrar filas de ejemplo (las que tenían relleno gris en el Excel original)
        if example_row_indices:
            # Ajustar índices: están basados en ws.iter_rows() (0-indexed en todo el sheet)
            # Necesitamos ajustar considerando que header_idx es relativo a rows
            data_rows = [
                r
                for idx, r in enumerate(data_rows)
                if (header_idx + 1 + idx) not in example_row_indices
            ]

        # Filtrar filas vacías
        data_rows = [r for r in data_rows if any(str(c).strip() for c in r)]

        return jsonify(
            {
                "headers": headers,
                "rows": data_rows,
                "total": len(data_rows),
                "tmp_path": tmp.name,
                "header_row_index": header_idx,
            }
        )
    except Exception as e:
        return jsonify({"error": f"Error leyendo Excel: {str(e)}"}), 500


@app.route("/api/preview", methods=["POST"])
def preview_cert():
    """Genera un PDF de vista previa con datos de la primera fila."""
    data = request.json
    plantilla_nombre = data.get("plantilla")
    fila = data.get("fila", {})
    config = data.get("config", {}) or {}

    # Mezclar config del agente (STPS, director) en la fila
    fila.update({k: v for k, v in config.items() if v})

    # Auto-generar folio si está vacío
    folio_raw = fila.get("folio", "")
    if not (folio_raw and str(folio_raw).strip()):
        if plantilla_nombre == "dc3":
            fila["folio"] = "DC3-001"
        elif "reconocimiento" in plantilla_nombre:
            fila["folio"] = "AG-001"
        elif "constancia" in plantilla_nombre:
            fila["folio"] = "CT-001"
        else:
            fila["folio"] = "001"

    plantilla_html = _render_plantilla(plantilla_nombre, fila)

    # Generar PDF temporal con Playwright
    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp_pdf.close()

    try:
        _generar_pdf_playwright(plantilla_html, tmp_pdf.name, single_page=True)
        return send_file(
            tmp_pdf.name, mimetype="application/pdf", download_name="preview.pdf"
        )
    except Exception as e:
        return jsonify({"error": f"Error generando preview: {str(e)}"}), 500


@app.route("/api/generate", methods=["POST"])
def generate_batch():
    """Inicia generación masiva en background."""
    global batch_state

    data = request.json
    plantilla_nombre = data.get("plantilla")
    tmp_excel_path = data.get("tmp_path")
    mapeo = data.get("mapeo", {})  # {campo_cert: columna_excel}
    filas = data.get("filas", [])
    config = data.get("config", {}) or {}  # {reg_stps, director_nombre, director_puesto}

    # Rango opcional: start_row, count
    start_row = data.get("start_row", 0)
    count = data.get("count", 0)  # 0 = todas desde start_row

    if not plantilla_nombre or not filas:
        return jsonify({"error": "Faltan datos: plantilla y filas"}), 400

    # Aplicar rango
    end_row = start_row + count if count > 0 else len(filas)
    end_row = min(end_row, len(filas))
    filas_range = filas[start_row:end_row]

    if not filas_range:
        return jsonify({"error": "El rango seleccionado no contiene filas"}), 400

    # Mezclar config del agente en cada fila (STPS, director, etc.)
    config_limpio = {k: v for k, v in config.items() if v}
    filas_con_config = [{**config_limpio, **fila} for fila in filas_range]

    # Reiniciar estado
    batch_state = {
        "status": "running",
        "total": len(filas_range),
        "completed": 0,
        "current_name": "",
        "started_at": datetime.now().isoformat(),
        "finished_at": None,
        "output_dir": "",
        "error": "",
        "failed_rows": [],
        "start_row": start_row,
    }

    # Thread en background
    t = threading.Thread(
        target=_run_batch, args=(plantilla_nombre, filas_con_config, mapeo, start_row), daemon=True
    )
    t.start()

    return jsonify({"status": "started", "total": len(filas_range)})


@app.route("/api/progress")
def progress():
    """SSE stream para progreso en tiempo real."""

    def generate():
        global batch_state
        while True:
            state = batch_state.copy()
            yield f"data: {json.dumps(state)}\n\n"
            if state["status"] in ("done", "error"):
                break
            time.sleep(0.5)

    return Response(stream_with_context(generate()), mimetype="text/event-stream")


@app.route("/api/download")
def download_zip():
    """Descarga todos los PDFs generados en un ZIP."""
    global batch_state
    output_dir = batch_state.get("output_dir", "")
    if not output_dir or not os.path.isdir(output_dir):
        return jsonify({"error": "No hay archivos para descargar"}), 400

    zip_path = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    zip_path.close()

    with zipfile.ZipFile(zip_path.name, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in sorted(Path(output_dir).glob("*.pdf")):
            zf.write(f, f.name)

    return send_file(
        zip_path.name,
        mimetype="application/zip",
        download_name=f"certificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
    )


@app.route("/api/open-folder")
def open_folder():
    """Abre la carpeta de output en el explorador de archivos."""
    global batch_state
    output_dir = batch_state.get("output_dir", "")
    if not output_dir or not os.path.isdir(output_dir):
        return jsonify({"error": "No hay carpeta"}), 400

    import platform
    import subprocess

    sysplat = platform.system()
    try:
        if sysplat == "Linux":
            subprocess.Popen(["xdg-open", output_dir])
        elif sysplat == "Windows":
            subprocess.Popen(["explorer", output_dir])
        elif sysplat == "Darwin":
            subprocess.Popen(["open", output_dir])
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Lógica interna ───


def _render_plantilla(nombre, datos):
    """Carga plantilla HTML y reemplaza placeholders {{campo}} con datos."""
    path = APP_ROOT / "plantillas" / f"{nombre}.html"
    if not path.exists():
        raise FileNotFoundError(f"Plantilla no encontrada: {nombre}")

    with open(path, "r", encoding="utf-8") as f:
        html = f.read()

    # Defaults opcionales (antes de reemplazar, para que los placeholders
    # como {{reg_stps}} siempre tengan un valor si vienen vacíos)
    datos["reg_stps"] = (datos.get("reg_stps") or "JUVH8204083R3-005").strip()
    # Defaults para firmas DC-3
    if nombre == "dc3":
        datos["firma_representante"] = datos.get("firma_representante") or "firmas/Firma_Soledad_Pastorutti.png"
        datos["firma_trabajadores"] = datos.get("firma_trabajadores") or ""

    # Reemplazar {{campo}} con valor correspondiente
    for key, val in datos.items():
        placeholder = "{{" + key + "}}"
        html = html.replace(placeholder, str(val) if val else "")

    # Expandir placeholders de caracteres individuales: {{curp_0}}..{{curp_17}}
    # Si existe {{curp_0}} en la plantilla y datos["curp"] tiene valor,
    # rellenar cada carácter automáticamente
    import re

    char_fields = {
        "curp": 18,  # CURP: 18 caracteres
        "rfc": 13,  # RFC: 13 caracteres
    }
    for base_field, length in char_fields.items():
        if base_field in datos and datos[base_field]:
            val = str(datos[base_field]).upper().ljust(length)[:length]
            for i in range(length):
                html = html.replace(f"{{{{{base_field}_{i}}}}}", val[i])
        else:
            for i in range(length):
                html = html.replace(f"{{{{{base_field}_{i}}}}}", "")

    # Fechas expandidas: si llega fecha_ini como "01/06/2026", parsear a componentes
    date_sources = [("fecha_ini", "fecha_ini"), ("fecha_fin", "fecha_fin")]
    for src_key, prefix in date_sources:
        if src_key in datos and datos[src_key] and "/" in str(datos[src_key]):
            parts = str(datos[src_key]).strip().split("/")
            if len(parts) == 3:
                dia, mes, año = parts
                # Expandir año
                for i, ch in enumerate(año.ljust(4)[:4]):
                    html = html.replace(f"{{{{{prefix}_año_{i}}}}}", ch)
                # Expandir mes
                for i, ch in enumerate(mes.ljust(2)[:2]):
                    html = html.replace(f"{{{{{prefix}_mes_{i}}}}}", ch)
                # Expandir día
                for i, ch in enumerate(dia.ljust(2)[:2]):
                    html = html.replace(f"{{{{{prefix}_dia_{i}}}}}", ch)

    # Fechas expandidas directas (si ya vienen separadas)
    date_groups = [
        ("fecha_ini_año", 4),
        ("fecha_ini_mes", 2),
        ("fecha_ini_dia", 2),
        ("fecha_fin_año", 4),
        ("fecha_fin_mes", 2),
        ("fecha_fin_dia", 2),
    ]
    for field, length in date_groups:
        if field in datos and datos[field]:
            val = str(datos[field]).ljust(length)[:length]
            for i in range(length):
                html = html.replace(f"{{{{{field}_{i}}}}}", val[i])
        else:
            for i in range(length):
                html = html.replace(f"{{{{{field}_{i}}}}}", "")

    # Reemplazar fecha actual si existe {{fecha_hoy}}
    html = html.replace("{{fecha_hoy}}", datetime.now().strftime("%d de %B de %Y"))

    # Reemplazar año actual
    html = html.replace("{{año}}", str(datetime.now().year))

    return html


def _generar_pdf_playwright(html_content, output_path, single_page=False):
    """Genera PDF desde HTML usando Playwright/Chromium."""
    import base64
    import mimetypes
    import re as re_mod
    import urllib.parse

    firmas_dir = APP_ROOT / "firmas"
    static_dir = APP_ROOT / "static"

    # Convertir <img src="firmas/foo.png"> a base64 inline
    def inline_image(match):
        full_tag = match.group(0)
        src_match = re_mod.search(r'src="([^"]+)"', full_tag)
        if not src_match:
            return full_tag
        src = src_match.group(1)
        # Determinar archivo físico
        if src.startswith("firmas/"):
            filepath = firmas_dir / src.replace("firmas/", "", 1)
        elif src.startswith("static/"):
            filepath = static_dir / src.replace("static/", "", 1)
        elif src.startswith("plantillas/"):
            filepath = APP_ROOT / src
        else:
            # Imagen suelta en la raíz del proyecto (ej. Gotita.png, AG_Principal.png)
            filepath = APP_ROOT / src
        if not filepath.exists():
            return full_tag
        # Convertir a base64
        try:
            data = filepath.read_bytes()
            mime, _ = mimetypes.guess_type(str(filepath))
            if not mime:
                ext = filepath.suffix.lower()
                mime = {
                    "png": "image/png",
                    "jpg": "image/jpeg",
                    "jpeg": "image/jpeg",
                    "gif": "image/gif",
                    "svg": "image/svg+xml",
                }.get(ext.lstrip("."), "image/png")
            b64 = base64.b64encode(data).decode("ascii")
            new_src = f"data:{mime};base64,{b64}"
            return full_tag.replace(f'src="{src}"', f'src="{new_src}"')
        except Exception as e:
            print(f"Error inlineando {src}: {e}")
            return full_tag

    html_content = re_mod.sub(r'<img[^>]*src="[^"]+"[^>]*>', inline_image, html_content)

    # Inlineado de fuentes en @font-face { src: url(...) } → data:font/ttf;base64,...
    # Necesario porque Playwright set_content usa origen data: y no puede cargar
    # recursos locales por path relativo. Soporta assets/ y rutas en raíz.
    def inline_font(match):
        full = match.group(0)
        url_match = re_mod.search(r"""url\(\s*['"]?([^'")]+)['"]?\s*\)""", full)
        if not url_match:
            return full
        url = url_match.group(1)
        # Solo inlinea paths locales con extensión de fuente
        if not re_mod.search(r"\.(ttf|otf|woff2?|eot)(\?|$)", url, re_mod.IGNORECASE):
            return full
        if url.startswith(("data:", "http://", "https://", "file://")):
            return full
        filepath = APP_ROOT / url
        if not filepath.exists():
            return full
        try:
            data = filepath.read_bytes()
            ext = filepath.suffix.lower().lstrip(".")
            mime = {"ttf": "font/ttf", "otf": "font/otf", "woff": "font/woff", "woff2": "font/woff2"}.get(ext, "font/ttf")
            b64 = base64.b64encode(data).decode("ascii")
            new_url = f"url(data:{mime};base64,{b64})"
            return full.replace(url, new_url)
        except Exception:
            return full

    html_content = re_mod.sub(r"url\([^)]+\)", inline_font, html_content)

    # Convertir href a file:// (links, no necesitan base64)
    firmas_url = "file://" + urllib.parse.quote(str(firmas_dir), safe="/:") + "/"
    static_url = "file://" + urllib.parse.quote(str(static_dir), safe="/:") + "/"
    html_content = html_content.replace('href="firmas/', f'href="{firmas_url}')
    html_content = html_content.replace('href="static/', f'href="{static_url}')

    # Usar chromium_headless_shell (más liviano) en vez de chromium completo
    # Buscamos el headless_shell dentro del bundle (extraído por PyInstaller)
    bundle_root = getattr(sys, "_MEIPASS", BASE_DIR)
    chrome_exec = None
    candidates = []

    # 1) Carpeta local del proyecto (build.spec pone el browser aquí)
    for sub in (Path(bundle_root) / "_ms-playwright").glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
        candidates.append(sub)
    for sub in (Path(bundle_root) / "_ms-playwright").glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
        candidates.append(sub)

    # 2) Ruta estándar de PyInstaller para Playwright (driver/.local-browsers)
    #    Esta es la que usa el bundled cuando playwright se importa sin env var
    for sub in (Path(bundle_root) / "playwright" / "driver" / "package" / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
        candidates.append(sub)
    for sub in (Path(bundle_root) / "playwright" / "driver" / "package" / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
        candidates.append(sub)

    # 3) Fallback .local-browsers en la raíz del bundle
    for sub in (Path(bundle_root) / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
        candidates.append(sub)
    for sub in (Path(bundle_root) / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
        candidates.append(sub)

    # 4) Cache global del sistema (~/.cache/ms-playwright) — dev mode
    for cache in [Path.home() / ".cache" / "ms-playwright"]:
        if cache.exists():
            for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
                candidates.append(sub)
            for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
                candidates.append(sub)

    # 5) Windows: %LOCALAPPDATA%\ms-playwright
    if os.name == "nt":
        local_app = os.environ.get("LOCALAPPDATA")
        if local_app:
            cache = Path(local_app) / "ms-playwright"
            if cache.exists():
                for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
                    candidates.append(sub)
                for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
                    candidates.append(sub)

    for c in candidates:
        if c.exists():
            chrome_exec = str(c)
            break

    with sync_playwright() as p:
        if chrome_exec:
            # Headless_shell: ~70 MB más liviano que el chromium completo
            browser = p.chromium.launch(
                executable_path=chrome_exec,
                args=["--allow-file-access-from-files"]
            )
        else:
            # Fallback: chromium completo (más pesado pero siempre presente)
            browser = p.chromium.launch(args=["--allow-file-access-from-files"])
        page = browser.new_page()
        page.set_content(html_content, wait_until="networkidle")

        is_landscape = "landscape" in html_content.lower()
        # Detectar si es DC-3 (usa Letter) vs A4
        page_format = "Letter" if "215.9mm" in html_content else "A4"

        page.pdf(
            path=output_path,
            format=page_format,
            landscape=is_landscape,
            print_background=True,
            margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
        )
        browser.close()


def _run_batch(plantilla_nombre, filas, mapeo, start_row=0):
    """Ejecuta la generación masiva en background.
    filas: lista de dicts con los datos ya mapeados {campo: valor}
    start_row: índice original (para reportar errores con el número real)
    """
    import traceback

    global batch_state

    output_dir = (
        RUNTIME_DIR / "output" / f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    batch_state["output_dir"] = str(output_dir)

    try:
        # ─── Buscar binario de chrome-headless-shell ───
        # Mismas ubicaciones que en _generar_pdf_playwright
        chrome_exec_batch = None
        cands = []
        for sub in (APP_ROOT / "_ms-playwright").glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
            cands.append(sub)
        for sub in (APP_ROOT / "_ms-playwright").glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
            cands.append(sub)
        for sub in (APP_ROOT / "playwright" / "driver" / "package" / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
            cands.append(sub)
        for sub in (APP_ROOT / "playwright" / "driver" / "package" / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
            cands.append(sub)
        for sub in (APP_ROOT / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
            cands.append(sub)
        for sub in (APP_ROOT / ".local-browsers").glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
            cands.append(sub)
        for cache in [Path.home() / ".cache" / "ms-playwright"]:
            if cache.exists():
                for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
                    cands.append(sub)
                for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
                    cands.append(sub)
        if os.name == "nt":
            local_app = os.environ.get("LOCALAPPDATA")
            if local_app:
                cache = Path(local_app) / "ms-playwright"
                if cache.exists():
                    for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-linux64/chrome-headless-shell"):
                        cands.append(sub)
                    for sub in cache.glob("chromium_headless_shell-*/chrome-headless-shell-win64/chrome-headless-shell.exe"):
                        cands.append(sub)
        for c in cands:
            if c.exists():
                chrome_exec_batch = str(c)
                break

        with sync_playwright() as p:
            if chrome_exec_batch:
                browser = p.chromium.launch(
                    executable_path=chrome_exec_batch,
                    args=["--allow-file-access-from-files"]
                )
            else:
                browser = p.chromium.launch(args=["--allow-file-access-from-files"])

            for i, datos in enumerate(filas):
                if batch_state["status"] == "error":
                    break

                # Los datos ya vienen mapeados desde el frontend
                nombre = datos.get("nombre", f"persona_{i + 1}")
                curso = datos.get("curso", "")
                folio_raw = datos.get("folio", "")
                folio = folio_raw.strip() if folio_raw else ""
                if not folio:
                    # Auto-generar folio según el tipo de plantilla
                    if plantilla_nombre == "dc3":
                        folio = f"DC3-{i + 1:03d}"
                    elif "reconocimiento" in plantilla_nombre:
                        folio = f"AG-{i + 1:03d}"
                    elif "constancia" in plantilla_nombre:
                        folio = f"CT-{i + 1:03d}"
                    else:
                        folio = f"{i + 1:03d}"
                    datos["folio"] = folio

                filename = _slugify(f"{nombre}_{curso}_{folio}") + ".pdf"
                filepath = output_dir / filename

                batch_state["current_name"] = nombre
                batch_state["completed"] = i

                try:
                    # Renderizar HTML
                    html = _render_plantilla(plantilla_nombre, datos)

                    # Convertir imágenes a base64 inline (firmas, logos)
                    import base64
                    import mimetypes
                    import re as re_mod_batch

                    firmas_dir_batch = APP_ROOT / "firmas"
                    static_dir_batch = APP_ROOT / "static"

                    def inline_image_batch(match):
                        full_tag = match.group(0)
                        src_match = re_mod_batch.search(r'src="([^"]+)"', full_tag)
                        if not src_match:
                            return full_tag
                        src = src_match.group(1)
                        if src.startswith("data:"):
                            return full_tag
                        if src.startswith("firmas/"):
                            filepath = firmas_dir_batch / src.replace("firmas/", "", 1)
                        elif src.startswith("static/"):
                            filepath = static_dir_batch / src.replace("static/", "", 1)
                        elif src.startswith("plantillas/"):
                            filepath = APP_ROOT / src
                        else:
                            filepath = APP_ROOT / src
                        if not filepath.exists():
                            return full_tag
                        try:
                            data = filepath.read_bytes()
                            mime, _ = mimetypes.guess_type(str(filepath))
                            if not mime:
                                ext = filepath.suffix.lower()
                                mime = {
                                    "png": "image/png",
                                    "jpg": "image/jpeg",
                                    "jpeg": "image/jpeg",
                                    "gif": "image/gif",
                                    "svg": "image/svg+xml",
                                }.get(ext.lstrip("."), "image/png")
                            b64 = base64.b64encode(data).decode("ascii")
                            new_src = f"data:{mime};base64,{b64}"
                            return full_tag.replace(f'src="{src}"', f'src="{new_src}"')
                        except:
                            return full_tag

                    html = re_mod_batch.sub(
                        r'<img[^>]*src="[^"]+"[^>]*>', inline_image_batch, html
                    )

                    def inline_font_batch(match):
                        full = match.group(0)
                        url_match = re_mod_batch.search(r"""url\(\s*['"]?([^'")]+)['"]?\s*\)""", full)
                        if not url_match:
                            return full
                        url = url_match.group(1)
                        if not re_mod_batch.search(r"\.(ttf|otf|woff2?|eot)(\?|$)", url, re_mod_batch.IGNORECASE):
                            return full
                        if url.startswith(("data:", "http://", "https://", "file://")):
                            return full
                        filepath = APP_ROOT / url
                        if not filepath.exists():
                            return full
                        try:
                            data = filepath.read_bytes()
                            ext = filepath.suffix.lower().lstrip(".")
                            mime = {"ttf": "font/ttf", "otf": "font/otf", "woff": "font/woff", "woff2": "font/woff2"}.get(ext, "font/ttf")
                            b64 = base64.b64encode(data).decode("ascii")
                            new_url = f"url(data:{mime};base64,{b64})"
                            return full.replace(url, new_url)
                        except:
                            return full

                    html = re_mod_batch.sub(r"url\([^)]+\)", inline_font_batch, html)

                    # Detectar orientación y formato
                    is_landscape = "landscape" in html.lower()
                    page_format = "Letter" if "215.9mm" in html else "A4"

                    # Generar PDF
                    page = browser.new_page()
                    page.set_content(html, wait_until="networkidle")
                    page.pdf(
                        path=str(filepath),
                        format=page_format,
                        landscape=is_landscape,
                        print_background=True,
                        margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
                    )
                    page.close()
                except Exception as e:
                    row_real = start_row + i + 1
                    batch_state.setdefault("failed_rows", []).append({
                        "row": row_real,
                        "nombre": nombre,
                        "error": str(e),
                    })
                    traceback.print_exc()

            browser.close()

        batch_state["completed"] = len(filas)
        batch_state["status"] = "done"
        batch_state["finished_at"] = datetime.now().isoformat()

    except Exception as e:
        batch_state["status"] = "error"
        batch_state["error"] = str(e)
        batch_state["finished_at"] = datetime.now().isoformat()
        traceback.print_exc()


def _slugify(text):
    """Limpia texto para usarlo como nombre de archivo."""
    import re

    text = text.strip().lower()
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"[\s-]+", "_", text)
    return text[:80]


# ─── Arranque ───


def main():
    url = f"http://{HOST}:{PORT}"
    print(f"\n{'=' * 50}")
    print(f"  🎓 Generador de Certificados AGASI")
    print(f"  Abriendo {url} ...")
    print(f"  Ctrl+C para cerrar")
    print(f"{'=' * 50}\n")

    # Abrir navegador automáticamente
    threading.Timer(0.5, lambda: webbrowser.open(url)).start()

    # En binario empaquetado (PyInstaller) no usar debug/reloader: el reloader
    # intenta reimportar el módulo principal y se confunde con sys.frozen.
    is_frozen = getattr(sys, "frozen", False)
    app.run(host=HOST, port=PORT, debug=not is_frozen, use_reloader=not is_frozen)


if __name__ == "__main__":
    main()
